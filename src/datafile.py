# -*- coding: utf-8 -*-
"""
    =============================
    Author: Linxson
    E-mail: 1853299@tongji.edu.cn
    =============================
"""
# 用来导入数据
from openpyxl import load_workbook


class Data:
    def __init__(self, dir):
        """dir为传入的文件名称
           e.g. r'D:\01A毕业设计\07图纸\管网\pyautocad\data_first.xlsx'
        """
        wb = load_workbook(dir)
        sheet_node = wb['Sheet1']
        self.nodeId = sheet_node['A']  # 节点ID
        self.nodex = sheet_node['I']  # 节点x坐标
        self.nodey = sheet_node['J']  # 节点y坐标
        self.nodeheight = sheet_node['B']  # 节点高程
        self.nodeq = sheet_node['E']  # 节点用水量
        self.nodehp = sheet_node['G']  # 节点总水头
        self.nodepre = sheet_node['H']  # 节点压力
        self.nodeTotal = len(self.nodeId)

        sheet_pipe = wb['Sheet2']
        self.pipeId = sheet_pipe['A']
        self.pipeTotal = len(self.pipeId)
        self.pipe_startnode = sheet_pipe['B']  # 上游节点ID
        self. pipe_endnode = sheet_pipe['C']  # 下游节点ID
        self.dn = sheet_pipe['D']  # 管径
        self.ln = sheet_pipe['E']  # 长度
        self.isclose = sheet_pipe['G']
        self.q = sheet_pipe['I']  # 流量
        self.absq = sheet_pipe['J']  # 流量绝对值
        self.vn = sheet_pipe['K']  # 获取流速
        self.unithj = sheet_pipe['L']  # 获取单位水头损失


class ConData:
    def __init__(self, dir, sheet):
        wb = load_workbook(dir)
        sheet_node = wb[sheet]  # sheet1为最高日最高时，2为消防，3为意外事故
        self.nodex = sheet_node['A']  # 节点x
        self.nodey = sheet_node['B']  # 节点y
        self.nodeTotal = len(self.nodex)
